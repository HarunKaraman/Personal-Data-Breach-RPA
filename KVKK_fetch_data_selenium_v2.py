# -*- coding: utf-8 -*-
"""
KVKK Fetch Data with Selenium Engine

August 2021

@project manager : Seha Solakoğlu
@author          : Harun Karaman

"""
import datetime
import os
import time
import openpyxl as excel
from selenium import webdriver
from pandas import DataFrame as df


def getwdandcwd():
    scriptcurrentdirectory = os.path.dirname(os.path.realpath(__file__))
    os.chdir(scriptcurrentdirectory)
    
def openexcelfile(excelfilename):
    global ws
    global wb
    wb = excel.load_workbook(excelfilename)
    ws = wb['Bildirimler']
    
def openbrowser_and_navigate_main_page(main_url):
    global browser
    try:
        browser.close()
    except:
        #print("Kapatılacak browser bulunmuyor.")
        print("")
    print("KVKK Veri İhlali Bildirimi - RPA Çalıştırıldı.")
    time.sleep(2)
    print("Webdriver Başlatılıyor!")
    browser = webdriver.Chrome('chromedriver_new.exe')
    browser.set_window_position(-10, 0)
    browser.set_window_size(720, 700)
    browser.get(main_url)
    #time.sleep(10)
    get_main_page_news()
    
def get_main_page_news():
    global news_title_1
    global news_title_2
    global news_title_3
    global news_title_4
    global news_title_5
    global news_title_6
    global news_title_7
    global news_title_8
    global news_title_9
    global news_title_10

    news_title_1 = browser.find_element_by_xpath('/html/body/div[3]/div/div[1]/div[1]/div[1]/div/div[2]/h3')
    news_title_2 = browser.find_element_by_xpath('/html/body/div[3]/div/div[1]/div[1]/div[2]/div/div[2]/h4/a')
    news_title_3 = browser.find_element_by_xpath('/html/body/div[3]/div/div[1]/div[1]/div[3]/div/div[2]/h4/a')
    news_title_4 = browser.find_element_by_xpath('/html/body/div[3]/div/div[1]/div[1]/div[4]/div/div[2]/h4/a')
    news_title_5 = browser.find_element_by_xpath('/html/body/div[3]/div/div[1]/div[1]/div[5]/div/div[2]/h4/a')
    news_title_6 = browser.find_element_by_xpath('/html/body/div[3]/div/div[1]/div[1]/div[6]/div/div[2]/h4/a')
    news_title_7 = browser.find_element_by_xpath('/html/body/div[3]/div/div[1]/div[1]/div[7]/div/div[2]/h4/a')
    news_title_8 = browser.find_element_by_xpath('/html/body/div[3]/div/div[1]/div[1]/div[8]/div/div[2]/h4/a')
    news_title_9 = browser.find_element_by_xpath('/html/body/div[3]/div/div[1]/div[1]/div[9]/div/div[2]/h4/a')
    news_title_10 = browser.find_element_by_xpath('/html/body/div[3]/div/div[1]/div[1]/div[10]/div/div[2]/h4/a')
    
    global news_release_date_1
    global news_release_date_2
    global news_release_date_3
    global news_release_date_4
    global news_release_date_5
    global news_release_date_6
    global news_release_date_7
    global news_release_date_8
    global news_release_date_9
    global news_release_date_10

    news_release_date_1 = browser.find_element_by_xpath('/html/body/div[3]/div/div[1]/div[1]/div[1]/div/div[2]/p[1]').text
    news_release_date_2 = browser.find_element_by_xpath('/html/body/div[3]/div/div[1]/div[1]/div[2]/div/div[2]/p').text
    news_release_date_3 = browser.find_element_by_xpath('/html/body/div[3]/div/div[1]/div[1]/div[3]/div/div[2]/p').text
    news_release_date_4 = browser.find_element_by_xpath('/html/body/div[3]/div/div[1]/div[1]/div[4]/div/div[2]/p').text
    news_release_date_5 = browser.find_element_by_xpath('/html/body/div[3]/div/div[1]/div[1]/div[5]/div/div[2]/p').text
    news_release_date_6 = browser.find_element_by_xpath('/html/body/div[3]/div/div[1]/div[1]/div[6]/div/div[2]/p').text
    news_release_date_7 = browser.find_element_by_xpath('/html/body/div[3]/div/div[1]/div[1]/div[7]/div/div[2]/p').text
    news_release_date_8 = browser.find_element_by_xpath('/html/body/div[3]/div/div[1]/div[1]/div[8]/div/div[2]/p').text
    news_release_date_9 = browser.find_element_by_xpath('/html/body/div[3]/div/div[1]/div[1]/div[9]/div/div[2]/p').text
    news_release_date_10 = browser.find_element_by_xpath('/html/body/div[3]/div/div[1]/div[1]/div[10]/div/div[2]/p').text

    global news_url_1
    global news_url_2
    global news_url_3
    global news_url_4
    global news_url_5
    global news_url_6
    global news_url_7
    global news_url_8
    global news_url_9
    global news_url_10

    news_url_1 = browser.find_element_by_xpath('/html/body/div[3]/div/div[1]/div[1]/div[1]/div/div[2]/div/a').get_attribute('href')
    news_url_2 = news_title_2.get_attribute('href')
    news_url_3 = news_title_3.get_attribute('href')
    news_url_4 = news_title_4.get_attribute('href')
    news_url_5 = news_title_5.get_attribute('href')
    news_url_6 = news_title_6.get_attribute('href')
    news_url_7 = news_title_7.get_attribute('href')
    news_url_8 = news_title_8.get_attribute('href')
    news_url_9 = news_title_9.get_attribute('href')
    news_url_10 = news_title_10.get_attribute('href')
    
    news_title_1 = browser.find_element_by_xpath('/html/body/div[3]/div/div[1]/div[1]/div[1]/div/div[2]/h3').text
    news_title_2 = browser.find_element_by_xpath('/html/body/div[3]/div/div[1]/div[1]/div[2]/div/div[2]/h4/a').text
    news_title_3 = browser.find_element_by_xpath('/html/body/div[3]/div/div[1]/div[1]/div[3]/div/div[2]/h4/a').text
    news_title_4 = browser.find_element_by_xpath('/html/body/div[3]/div/div[1]/div[1]/div[4]/div/div[2]/h4/a').text
    news_title_5 = browser.find_element_by_xpath('/html/body/div[3]/div/div[1]/div[1]/div[5]/div/div[2]/h4/a').text
    news_title_6 = browser.find_element_by_xpath('/html/body/div[3]/div/div[1]/div[1]/div[6]/div/div[2]/h4/a').text
    news_title_7 = browser.find_element_by_xpath('/html/body/div[3]/div/div[1]/div[1]/div[7]/div/div[2]/h4/a').text
    news_title_8 = browser.find_element_by_xpath('/html/body/div[3]/div/div[1]/div[1]/div[8]/div/div[2]/h4/a').text
    news_title_9 = browser.find_element_by_xpath('/html/body/div[3]/div/div[1]/div[1]/div[9]/div/div[2]/h4/a').text
    news_title_10 = browser.find_element_by_xpath('/html/body/div[3]/div/div[1]/div[1]/div[10]/div/div[2]/h4/a').text
    
    
def compare_mainpage_news_with_excel_data():
    last_empty_row = len(list(ws.rows))
    #print(last_empty_row)
    #char = get_column_letter(2)
    son_bildirim_tarihi = ws["B"+str(last_empty_row)].value
    son_bildirim_baslik = ws["C"+str(last_empty_row)].value
    son_bildirim_link = ws["C"+str(last_empty_row)].value
    
    if (son_bildirim_tarihi == news_release_date_1 and 
        son_bildirim_baslik == news_title_1):
        try:
            print("data güncel")
        except:
            print("1.haberin kontrolünde hata alındı.")
    elif (son_bildirim_tarihi == news_release_date_10 and 
        son_bildirim_baslik == news_title_10):
        
        try:
            last_empty_row = str(last_empty_row+1)
            timestamp = datetime.datetime.now().strftime("%d-%m-%Y %H:%M:%S")
            browser.get(news_url_9)
            print("9.Veri İhlali Bildiriminin yer aldığı sayfa açılıyor.")
            browser.execute_script("window.scrollTo(0,1150);")
            try:
                title = browser.find_element_by_xpath('/html/body/div[3]/div/div[1]/div/div/div/div[2]/h3').text
                body = browser.find_element_by_xpath('/html/body/div[3]/div/div[1]/div/div/div/div[2]/div').text
            except:
                print("ERROR")
                title = "ERROR"
                body = "ERROR"
            guncel_flag = 1
            ws["A"+last_empty_row].value = timestamp
            ws["B"+last_empty_row].value = news_release_date_9
            ws["C"+last_empty_row].value = news_title_9
            ws["D"+last_empty_row].value = news_url_9
            ws["E"+last_empty_row].value = title
            ws["F"+last_empty_row].value = body
            ws["G"+last_empty_row].value = guncel_flag
            time.sleep(2)    
            print("9.Veri İhlali Bildirimi İçeriği önbelleğe aktarıldı.")
        except:
            print("9.Haberin kaydedilmesinde hata alındı.")
    
    elif (son_bildirim_tarihi == news_release_date_9 and 
        son_bildirim_baslik == news_title_9):
        
        try:
            last_empty_row = str(last_empty_row+1)
            timestamp = datetime.datetime.now().strftime("%d-%m-%Y %H:%M:%S")
            browser.get(news_url_8)
            print("8.Veri İhlali Bildiriminin yer aldığı sayfa açılıyor.")
            browser.execute_script("window.scrollTo(0,1150);")
            try:
                title = browser.find_element_by_xpath('/html/body/div[3]/div/div[1]/div/div/div/div[2]/h3').text
                body = browser.find_element_by_xpath('/html/body/div[3]/div/div[1]/div/div/div/div[2]/div').text
            except:
                print("ERROR")
                title = "ERROR"
                body = "ERROR"
            guncel_flag = 1
            ws["A"+last_empty_row].value = timestamp
            ws["B"+last_empty_row].value = news_release_date_8
            ws["C"+last_empty_row].value = news_title_8
            ws["D"+last_empty_row].value = news_url_8
            ws["E"+last_empty_row].value = title
            ws["F"+last_empty_row].value = body
            ws["G"+last_empty_row].value = guncel_flag
            time.sleep(2)    
            print("8.Veri İhlali Bildirimi İçeriği önbelleğe aktarıldı.")
        except:
            print("8.Haberin kaydedilmesinde hata alındı.")
    
    elif (son_bildirim_tarihi == news_release_date_8 and 
        son_bildirim_baslik == news_title_8):
        
        try:
            last_empty_row = str(last_empty_row+1)
            timestamp = datetime.datetime.now().strftime("%d-%m-%Y %H:%M:%S")
            browser.get(news_url_7)
            print("7.Veri İhlali Bildiriminin yer aldığı sayfa açılıyor.")
            browser.execute_script("window.scrollTo(0,1150);")
            try:
                title = browser.find_element_by_xpath('/html/body/div[3]/div/div[1]/div/div/div/div[2]/h3').text
                body = browser.find_element_by_xpath('/html/body/div[3]/div/div[1]/div/div/div/div[2]/div').text
            except:
                print("ERROR")
                title = "ERROR"
                body = "ERROR"
            guncel_flag = 1
            ws["A"+last_empty_row].value = timestamp
            ws["B"+last_empty_row].value = news_release_date_7
            ws["C"+last_empty_row].value = news_title_7
            ws["D"+last_empty_row].value = news_url_7
            ws["E"+last_empty_row].value = title
            ws["F"+last_empty_row].value = body
            ws["G"+last_empty_row].value = guncel_flag
            time.sleep(2)    
            print("7.Veri İhlali Bildirimi İçeriği önbelleğe aktarıldı.")
        except:
            print("7.Haberin kaydedilmesinde hata alındı.")
    
    elif (son_bildirim_tarihi == news_release_date_7 and 
        son_bildirim_baslik == news_title_7):
        
        try:
            last_empty_row = str(last_empty_row+1)
            timestamp = datetime.datetime.now().strftime("%d-%m-%Y %H:%M:%S")
            browser.get(news_url_6)
            print("6.Veri İhlali Bildiriminin yer aldığı sayfa açılıyor.")
            browser.execute_script("window.scrollTo(0,1150);")
            try:
                title = browser.find_element_by_xpath('/html/body/div[3]/div/div[1]/div/div/div/div[2]/h3').text
                body = browser.find_element_by_xpath('/html/body/div[3]/div/div[1]/div/div/div/div[2]/div').text
            except:
                print("ERROR")
                title = "ERROR"
                body = "ERROR"
            guncel_flag = 1
            ws["A"+last_empty_row].value = timestamp
            ws["B"+last_empty_row].value = news_release_date_6
            ws["C"+last_empty_row].value = news_title_6
            ws["D"+last_empty_row].value = news_url_6
            ws["E"+last_empty_row].value = title
            ws["F"+last_empty_row].value = body
            ws["G"+last_empty_row].value = guncel_flag
            time.sleep(2)    
            print("6.Veri İhlali Bildirimi İçeriği önbelleğe aktarıldı.")
        except:
            print("6.Haberin kaydedilmesinde hata alındı.")
    
    elif (son_bildirim_tarihi == news_release_date_6 and 
        son_bildirim_baslik == news_title_6):
        
        try:
            last_empty_row = str(last_empty_row+1)
            timestamp = datetime.datetime.now().strftime("%d-%m-%Y %H:%M:%S")
            browser.get(news_url_5)
            print("5.Veri İhlali Bildiriminin yer aldığı sayfa açılıyor.")
            browser.execute_script("window.scrollTo(0,1150);")
            try:
                title = browser.find_element_by_xpath('/html/body/div[3]/div/div[1]/div/div/div/div[2]/h3').text
                body = browser.find_element_by_xpath('/html/body/div[3]/div/div[1]/div/div/div/div[2]/div').text
            except:
                print("ERROR")
                title = "ERROR"
                body = "ERROR"
            guncel_flag = 1
            ws["A"+last_empty_row].value = timestamp
            ws["B"+last_empty_row].value = news_release_date_5
            ws["C"+last_empty_row].value = news_title_5
            ws["D"+last_empty_row].value = news_url_5
            ws["E"+last_empty_row].value = title
            ws["F"+last_empty_row].value = body
            ws["G"+last_empty_row].value = guncel_flag
            time.sleep(2)    
            print("5.Veri İhlali Bildirimi İçeriği önbelleğe aktarıldı.")
        except:
            print("5.Haberin kaydedilmesinde hata alındı.")
    
    elif (son_bildirim_tarihi == news_release_date_5 and 
        son_bildirim_baslik == news_title_5):
        
        try:
            last_empty_row = str(last_empty_row+1)
            timestamp = datetime.datetime.now().strftime("%d-%m-%Y %H:%M:%S")
            browser.get(news_url_4)
            print("4.Veri İhlali Bildiriminin yer aldığı sayfa açılıyor.")
            browser.execute_script("window.scrollTo(0,1150);")
            try:
                title = browser.find_element_by_xpath('/html/body/div[3]/div/div[1]/div/div/div/div[2]/h3').text
                body = browser.find_element_by_xpath('/html/body/div[3]/div/div[1]/div/div/div/div[2]/div').text
            except:
                print("ERROR")
                title = "ERROR"
                body = "ERROR"
            guncel_flag = 1
            ws["A"+last_empty_row].value = timestamp
            ws["B"+last_empty_row].value = news_release_date_4
            ws["C"+last_empty_row].value = news_title_4
            ws["D"+last_empty_row].value = news_url_4
            ws["E"+last_empty_row].value = title
            ws["F"+last_empty_row].value = body
            ws["G"+last_empty_row].value = guncel_flag
            time.sleep(2)    
            print("4.Veri İhlali Bildirimi İçeriği önbelleğe aktarıldı.")
        except:
            print("4.Haberin kaydedilmesinde hata alındı.")
    
    elif (son_bildirim_tarihi == news_release_date_4 and 
        son_bildirim_baslik == news_title_4):
        
        try:
            last_empty_row = str(last_empty_row+1)
            timestamp = datetime.datetime.now().strftime("%d-%m-%Y %H:%M:%S")
            browser.get(news_url_3)
            print("3.Veri İhlali Bildiriminin yer aldığı sayfa açılıyor.")
            browser.execute_script("window.scrollTo(0,1150);")
            try:
                title = browser.find_element_by_xpath('/html/body/div[3]/div/div[1]/div/div/div/div[2]/h3').text
                body = browser.find_element_by_xpath('/html/body/div[3]/div/div[1]/div/div/div/div[2]/div').text
            except:
                print("ERROR")
                title = "ERROR"
                body = "ERROR"
            guncel_flag = 1
            ws["A"+last_empty_row].value = timestamp
            ws["B"+last_empty_row].value = news_release_date_3
            ws["C"+last_empty_row].value = news_title_3
            ws["D"+last_empty_row].value = news_url_3
            ws["E"+last_empty_row].value = title
            ws["F"+last_empty_row].value = body
            ws["G"+last_empty_row].value = guncel_flag
            time.sleep(2)    
            print("3.Veri İhlali Bildirimi İçeriği önbelleğe aktarıldı.")
        except:
            print("3.Haberin kaydedilmesinde hata alındı.")
    
    elif (son_bildirim_tarihi == news_release_date_3 and 
        son_bildirim_baslik == news_title_3):
        
        try:
            last_empty_row = str(last_empty_row+1)
            timestamp = datetime.datetime.now().strftime("%d-%m-%Y %H:%M:%S")
            browser.get(news_url_2)
            print("2.Veri İhlali Bildiriminin yer aldığı sayfa açılıyor.")
            browser.execute_script("window.scrollTo(0,1150);")
            try:
                title = browser.find_element_by_xpath('/html/body/div[3]/div/div[1]/div/div/div/div[2]/h3').text
                body = browser.find_element_by_xpath('/html/body/div[3]/div/div[1]/div/div/div/div[2]/div').text
            except:
                print("ERROR")
                title = "ERROR"
                body = "ERROR"
            guncel_flag = 1
            ws["A"+last_empty_row].value = timestamp
            ws["B"+last_empty_row].value = news_release_date_2
            ws["C"+last_empty_row].value = news_title_2
            ws["D"+last_empty_row].value = news_url_2
            ws["E"+last_empty_row].value = title
            ws["F"+last_empty_row].value = body
            ws["G"+last_empty_row].value = guncel_flag
            time.sleep(2)    
            print("2.Veri İhlali Bildirimi İçeriği önbelleğe aktarıldı.")
        except:
            print("2.Haberin kaydedilmesinde hata alındı.")
    
    elif (son_bildirim_tarihi == news_release_date_2 and 
        son_bildirim_baslik == news_title_2):
        
        try:
            last_empty_row = str(last_empty_row+1)
            timestamp = datetime.datetime.now().strftime("%d-%m-%Y %H:%M:%S")
            browser.get(news_url_1)
            print("1.Veri İhlali Bildiriminin yer aldığı sayfa açılıyor.")
            browser.execute_script("window.scrollTo(0,850);")
            try:
                title = browser.find_element_by_xpath('/html/body/div[3]/div/div[1]/div/div/div/div[2]/h3').text
                body = browser.find_element_by_xpath('/html/body/div[3]/div/div[1]/div/div/div/div[2]/div').text
            except:
                print("ERROR")
                title = "ERROR"
                body = "ERROR"
            guncel_flag = 1
            ws["A"+last_empty_row].value = timestamp
            ws["B"+last_empty_row].value = news_release_date_1
            ws["C"+last_empty_row].value = news_title_1
            ws["D"+last_empty_row].value = news_url_1
            ws["E"+last_empty_row].value = title
            ws["F"+last_empty_row].value = body
            ws["G"+last_empty_row].value = guncel_flag
            time.sleep(2)    
            print("1.Veri İhlali Bildirimi İçeriği önbelleğe aktarıldı.")
        except:
            print("1.Haberin kaydedilmesinde hata alındı.")
    else:
        
        try:
            last_empty_row = str(last_empty_row+1)
            timestamp = datetime.datetime.now().strftime("%d-%m-%Y %H:%M:%S")
            browser.get(news_url_10)
            print("10.Veri İhlali Bildiriminin yer aldığı sayfa açılıyor.")
            browser.execute_script("window.scrollTo(0,1150);")
            try:
                title = browser.find_element_by_xpath('/html/body/div[3]/div/div[1]/div/div/div/div[2]/h3').text
                body = browser.find_element_by_xpath('/html/body/div[3]/div/div[1]/div/div/div/div[2]/div').text
            except:
                print("ERROR")
                title = "ERROR"
                body = "ERROR"            
            guncel_flag = 1
            ws["A"+last_empty_row].value = timestamp
            ws["B"+last_empty_row].value = news_release_date_10
            ws["C"+last_empty_row].value = news_title_10
            ws["D"+last_empty_row].value = news_url_10
            ws["E"+last_empty_row].value = title
            ws["F"+last_empty_row].value = body
            ws["G"+last_empty_row].value = guncel_flag
            time.sleep(2)    
            print("10.Veri İhlali Bildirimi İçeriği önbelleğe aktarıldı.")
        except:
            print("10.Haberin kaydedilmesinde hata alındı.")
    time.sleep(1)
    
    #Çıktılar kaydedilir.
    wb.save("KVKK_RESULTS.xlsx")

kvkklink10 = 'https://www.kvkk.gov.tr/veri-ihlali-bildirimi/?&page=10'
kvkklink9 = 'https://www.kvkk.gov.tr/veri-ihlali-bildirimi/?&page=9'
kvkklink8 = 'https://www.kvkk.gov.tr/veri-ihlali-bildirimi/?&page=8'
kvkklink7 = 'https://www.kvkk.gov.tr/veri-ihlali-bildirimi/?&page=7'
kvkklink6 = 'https://www.kvkk.gov.tr/veri-ihlali-bildirimi/?&page=6'
kvkklink5 = 'https://www.kvkk.gov.tr/veri-ihlali-bildirimi/?&page=5'
kvkklink4 = 'https://www.kvkk.gov.tr/veri-ihlali-bildirimi/?&page=4'
kvkklink3 = 'https://www.kvkk.gov.tr/veri-ihlali-bildirimi/?&page=3'
kvkklink2 = 'https://www.kvkk.gov.tr/veri-ihlali-bildirimi/?&page=2'
kvkklink1 = 'https://www.kvkk.gov.tr/veri-ihlali-bildirimi/?&page=1'


def start():
    openexcelfile("KVKK_RESULTS.xlsx")
    getwdandcwd()
    openbrowser_and_navigate_main_page(kvkklink10)
    for i in range(1,11):
        compare_mainpage_news_with_excel_data()
    openbrowser_and_navigate_main_page(kvkklink9)
    for i in range(1,11):
        compare_mainpage_news_with_excel_data()
    openbrowser_and_navigate_main_page(kvkklink8)
    for i in range(1,11):
        compare_mainpage_news_with_excel_data()
    openbrowser_and_navigate_main_page(kvkklink7)
    for i in range(1,11):
        compare_mainpage_news_with_excel_data()
    openbrowser_and_navigate_main_page(kvkklink6)
    for i in range(1,11):
        compare_mainpage_news_with_excel_data()
    openbrowser_and_navigate_main_page(kvkklink5)
    for i in range(1,11):
        compare_mainpage_news_with_excel_data()
    openbrowser_and_navigate_main_page(kvkklink4)
    for i in range(1,11):
        compare_mainpage_news_with_excel_data()
    openbrowser_and_navigate_main_page(kvkklink3)
    for i in range(1,11):
        compare_mainpage_news_with_excel_data()
    openbrowser_and_navigate_main_page(kvkklink2)
    for i in range(1,11):
        compare_mainpage_news_with_excel_data()
    openbrowser_and_navigate_main_page(kvkklink1)
    for i in range(1,11):
        compare_mainpage_news_with_excel_data()
    print("Önbellekteki veriler EXCEL dosyasına aktarıldı.")

if __name__ == "__main__":
    start()